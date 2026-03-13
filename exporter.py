# -*- coding: utf-8 -*-
"""
TapTap评价分析系统 - Excel导出模块 (优化版)

优化点：
1. 添加新字段：游戏时长、多分类、评价摘要
2. 添加分天趋势统计
3. 添加饼图

@author: TapTap Analyzer
@version: 1.1.0
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Any
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

from config import OUTPUT_CONFIG, LOG_CONFIG

logging.basicConfig(
    level=getattr(logging, LOG_CONFIG['level']),
    format=LOG_CONFIG['format'],
    datefmt=LOG_CONFIG['date_format']
)
logger = logging.getLogger(__name__)


def summarize_text(text: str, max_length: int = 50) -> str:
    """
    提取文本核心要点摘要
    
    Args:
        text: 原始文本
        max_length: 最大长度
        
    Returns:
        摘要文本
    """
    if not text:
        return ""
    
    # 关键句提取规则
    sentences = text.replace('。', '。\n').replace('！', '！\n').replace('？', '？\n').split('\n')
    sentences = [s.strip() for s in sentences if s.strip()]
    
    # 优先级关键词
    priority_keywords = [
        '建议', '希望', '应该', '需要', '问题', '缺点', '优点',
        '太', '很', '非常', '严重', '太多', '太少',
        '好玩', '垃圾', '卡', '闪退', '氪金', '外挂', '匹配'
    ]
    
    # 评分每句话的重要性
    scored_sentences = []
    for sent in sentences:
        score = 0
        for kw in priority_keywords:
            if kw in sent:
                score += 1
        # 长度适中加分
        if 10 < len(sent) < 50:
            score += 1
        scored_sentences.append((sent, score))
    
    # 按分数排序，取最重要的句子
    scored_sentences.sort(key=lambda x: -x[1])
    
    summary = ""
    for sent, score in scored_sentences:
        if len(summary) + len(sent) <= max_length:
            if summary:
                summary += "；"
            summary += sent[:max_length - len(summary)]
        else:
            break
    
    if not summary:
        summary = text[:max_length]
    
    return summary


class ExcelExporter:
    """
    Excel导出器类
    """
    
    def __init__(self):
        self.output_dir = OUTPUT_CONFIG['output_dir']
        os.makedirs(self.output_dir, exist_ok=True)
        
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font_white = Font(bold=True, size=11, color='FFFFFF')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export(
        self,
        reviews: List[Dict[str, Any]],
        stats: Dict[str, Any],
        game_name: str = "游戏"
    ) -> str:
        """
        导出评价数据到Excel
        
        Args:
            reviews: 评价列表
            stats: 统计数据
            game_name: 游戏名称
            
        Returns:
            Excel文件路径
        """
        timestamp = datetime.now().strftime('%Y%m%d')
        filename = f"{game_name}_评价分析_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: 评价明细
        ws1 = wb.active
        ws1.title = "评价明细"
        self._write_reviews_sheet(ws1, reviews)
        
        # Sheet 2: 分类统计
        ws2 = wb.create_sheet("分类统计")
        self._write_category_stats(ws2, stats, reviews)
        
        # Sheet 3: 分天趋势
        ws3 = wb.create_sheet("分天趋势")
        self._write_daily_trend(ws3, reviews)
        
        # Sheet 4: 情感分布饼图
        ws4 = wb.create_sheet("情感分布")
        self._write_sentiment_pie(ws4, reviews)
        
        wb.save(filepath)
        logger.info(f"Excel文件已保存: {filepath}")
        
        return filepath
    
    def _write_reviews_sheet(self, ws, reviews: List[Dict[str, Any]]):
        """
        写入评价明细Sheet
        """
        headers = [
            '序号', '评价日期', '用户名', '星级', '游戏时长', '设备',
            '评价内容', '内容摘要', '情感分数', '情感倾向',
            '主分类', '副分类', '所有分类', '匹配关键词'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        for row, review in enumerate(reviews, 2):
            # 序号
            ws.cell(row=row, column=1, value=row-1).border = self.border
            
            # 评价日期
            ws.cell(row=row, column=2, value=review.get('review_date', '')).border = self.border
            
            # 用户名
            ws.cell(row=row, column=3, value=review.get('user_name', '匿名')).border = self.border
            
            # 星级
            rating = review.get('rating', 0)
            ws.cell(row=row, column=4, value=f"{rating}星" if rating else '').border = self.border
            
            # 游戏时长
            play_time_str = review.get('play_time_str', '')
            if not play_time_str:
                play_time = review.get('play_time', 0)
                if play_time:
                    play_time_str = f"{play_time/60:.1f}小时"
            ws.cell(row=row, column=5, value=play_time_str).border = self.border
            
            # 设备
            ws.cell(row=row, column=6, value=review.get('device', '')).border = self.border
            
            # 评价内容
            content = review.get('content', '')
            ws.cell(row=row, column=7, value=content).border = self.border
            
            # 内容摘要
            summary = summarize_text(content, 60)
            ws.cell(row=row, column=8, value=summary).border = self.border
            
            # 情感分数
            sentiment_score = review.get('sentiment_score', 0.5)
            ws.cell(row=row, column=9, value=round(sentiment_score, 3)).border = self.border
            
            # 情感倾向
            ws.cell(row=row, column=10, value=review.get('sentiment_label', '中性')).border = self.border
            
            # 主分类
            ws.cell(row=row, column=11, value=review.get('primary_category', '其他')).border = self.border
            
            # 副分类
            ws.cell(row=row, column=12, value=review.get('secondary_category', '其他')).border = self.border
            
            # 所有分类
            all_cats = review.get('all_primary_categories', [])
            all_cats_str = '、'.join(all_cats) if all_cats else ''
            ws.cell(row=row, column=13, value=all_cats_str).border = self.border
            
            # 匹配关键词
            matched = review.get('matched_keywords', [])
            if isinstance(matched, list) and matched:
                if isinstance(matched[0], dict):
                    keywords_str = '、'.join([m.get('keyword', '') for m in matched])
                else:
                    keywords_str = '、'.join(matched)
            else:
                keywords_str = ''
            ws.cell(row=row, column=14, value=keywords_str).border = self.border
        
        # 调整列宽
        column_widths = [6, 12, 12, 6, 10, 12, 60, 40, 10, 8, 12, 15, 25, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        logger.info(f"评价明细Sheet创建完成，共 {len(reviews)} 条记录")
    
    def _write_category_stats(self, ws, stats: Dict[str, Any], reviews: List[Dict[str, Any]]):
        """
        写入分类统计Sheet
        """
        # 一级分类统计
        ws.cell(row=1, column=1, value="一级分类统计").font = self.header_font
        ws.merge_cells('A1:C1')
        
        ws.cell(row=2, column=1, value="分类").font = self.header_font
        ws.cell(row=2, column=2, value="数量").font = self.header_font
        ws.cell(row=2, column=3, value="占比").font = self.header_font
        
        primary_dist = stats.get('primary_distribution', {})
        total = stats.get('total_count', 1)
        
        row = 3
        for cat, count in sorted(primary_dist.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%")
            row += 1
        
        # 多标签分类统计（所有匹配的分类）
        row += 2
        ws.cell(row=row, column=1, value="多标签分类统计（含所有匹配）").font = self.header_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        ws.cell(row=row, column=1, value="分类").font = self.header_font
        ws.cell(row=row, column=2, value="出现次数").font = self.header_font
        ws.cell(row=row, column=3, value="备注").font = self.header_font
        row += 1
        
        all_primaries_dist = stats.get('all_primaries_distribution', {})
        for cat, count in sorted(all_primaries_dist.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value="一条评价可能属于多个分类")
            row += 1
        
        # 情感分布统计
        row += 2
        ws.cell(row=row, column=1, value="情感分布统计").font = self.header_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        ws.cell(row=row, column=1, value="情感").font = self.header_font
        ws.cell(row=row, column=2, value="数量").font = self.header_font
        ws.cell(row=row, column=3, value="占比").font = self.header_font
        row += 1
        
        sentiment_dist = {'正向': 0, '中性': 0, '负向': 0}
        for r in reviews:
            label = r.get('sentiment_label', '中性')
            if label in sentiment_dist:
                sentiment_dist[label] += 1
        
        for label, count in sentiment_dist.items():
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%")
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        
        logger.info("分类统计Sheet创建完成")
    
    def _write_daily_trend(self, ws, reviews: List[Dict[str, Any]]):
        """
        写入分天趋势Sheet
        """
        # 统计每天的情感分布
        daily_stats = defaultdict(lambda: {'正向': 0, '中性': 0, '负向': 0, 'total': 0})
        
        for review in reviews:
            date = review.get('review_date', '')
            if date:
                label = review.get('sentiment_label', '中性')
                daily_stats[date][label] += 1
                daily_stats[date]['total'] += 1
        
        # 按日期排序
        sorted_dates = sorted(daily_stats.keys())
        
        # 写入表头
        headers = ['日期', '正向', '中性', '负向', '总计', '正向占比', '负向占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.border
        
        # 写入数据
        for row, date in enumerate(sorted_dates, 2):
            stats = daily_stats[date]
            total = stats['total']
            
            ws.cell(row=row, column=1, value=date).border = self.border
            ws.cell(row=row, column=2, value=stats['正向']).border = self.border
            ws.cell(row=row, column=3, value=stats['中性']).border = self.border
            ws.cell(row=row, column=4, value=stats['负向']).border = self.border
            ws.cell(row=row, column=5, value=total).border = self.border
            ws.cell(row=row, column=6, value=f"{stats['正向']/total*100:.1f}%" if total else '0%').border = self.border
            ws.cell(row=row, column=7, value=f"{stats['负向']/total*100:.1f}%" if total else '0%').border = self.border
        
        # 添加趋势图
        if len(sorted_dates) > 1:
            # 条数趋势图
            chart1 = BarChart()
            chart1.type = "col"
            chart1.title = "每日评价条数趋势"
            chart1.y_axis.title = "条数"
            chart1.x_axis.title = "日期"
            
            data = Reference(ws, min_col=5, min_row=1, max_row=len(sorted_dates)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_dates)+1)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.shape = 4
            ws.add_chart(chart1, "I2")
        
        # 调整列宽
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        logger.info("分天趋势Sheet创建完成")
    
    def _write_sentiment_pie(self, ws, reviews: List[Dict[str, Any]]):
        """
        写入情感分布饼图Sheet
        """
        # 统计情感分布
        sentiment_dist = {'正向': 0, '中性': 0, '负向': 0}
        for review in reviews:
            label = review.get('sentiment_label', '中性')
            if label in sentiment_dist:
                sentiment_dist[label] += 1
        
        # 写入数据
        ws.cell(row=1, column=1, value="情感").font = self.header_font
        ws.cell(row=1, column=2, value="数量").font = self.header_font
        
        row = 2
        for label, count in sentiment_dist.items():
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # 创建饼图
        pie = PieChart()
        pie.title = "情感分布饼图"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=4)
        cats = Reference(ws, min_col=1, min_row=2, max_row=4)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        
        # 显示百分比标签
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False
        
        ws.add_chart(pie, "D2")
        
        # 添加星级分布
        rating_dist = defaultdict(int)
        for review in reviews:
            rating = review.get('rating', 0)
            if rating:
                rating_dist[rating] += 1
        
        ws.cell(row=1, column=4, value="星级").font = self.header_font
        ws.cell(row=1, column=5, value="数量").font = self.header_font
        
        row = 2
        for rating in range(5, 0, -1):
            ws.cell(row=row, column=4, value=f"{rating}星")
            ws.cell(row=row, column=5, value=rating_dist.get(rating, 0))
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        
        logger.info("情感分布Sheet创建完成")


def export_reviews(
    reviews: List[Dict[str, Any]],
    stats: Dict[str, Any],
    game_name: str = "游戏"
) -> str:
    """
    便捷函数：导出评价数据
    """
    exporter = ExcelExporter()
    return exporter.export(reviews, stats, game_name)


if __name__ == '__main__':
    # 测试
    test_reviews = [
        {
            'review_date': '2024-02-15',
            'user_name': '测试用户',
            'rating': 4,
            'play_time_str': '3.5',
            'device': 'iPhone 14',
            'content': '游戏很好玩，但是有些卡顿，希望优化一下。',
            'sentiment_score': 0.65,
            'sentiment_label': '正向',
            'primary_category': '技术问题',
            'secondary_category': '卡顿/掉帧',
            'all_primary_categories': ['技术问题', '正向评价'],
            'matched_keywords': [{'keyword': '卡顿', 'primary': '技术问题', 'secondary': '卡顿/掉帧'}]
        }
    ]
    
    test_stats = {
        'primary_distribution': {'技术问题': 1},
        'secondary_distribution': {'技术问题': {'卡顿/掉帧': 1}},
        'all_primaries_distribution': {'技术问题': 1, '正向评价': 1},
        'total_count': 1
    }
    
    filepath = export_reviews(test_reviews, test_stats, "测试游戏")
    print(f"文件已保存: {filepath}")
